# Importacion de pandas para la lectura de archivos Excel
import pandas as pd
from openpyxl import load_workbook
import os

URL_PATH_EXCEL = r'rendimiento.xlsx'
columnas = [
    "Código", "Fecha", "Mina", "Peso Tolva [Kg]", "Ley Cabeza [gr/ton]",
    "Peso Ingresa [gr/ton]", "Peso Rech. [Kg]", "Ley RV [gr/ton]", 
    "Ley RC [gr/ton]", "Contenido Au miligramos", "Peso Sel. Veta [Kg]", 
    "Ley SV [gr/ton]", "Peso Sel. Caja [Kg]", "Ley SC [gr/ton]", 
    "Contenido Au Caja [Mg]", "Peso Veta Total [Kg]", "Peso Caja Total [Kg]", 
    "Reducción Masa [%]", "Rel. % Veta Post", "Rel. % Caja Post", 
    "Ley Rechazo Calc [gr/ton]", "Ley Selectado Calc [gr/ton]", 
    "Recuperación Au Retenido [Kg]", "Recuperación Au Equipo [Kg]", 
    "Recuperación Au Rechazo [Kg]", "Recuperación Au Veta [Kg]", 
    "Modelo", "Confianza", "Num Pasadas"
]

def read_xlss_excel(archivo):
    wb = load_workbook(archivo)
    #print(wb.sheetnames)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        print(row)
    columna_b = [celda.value for celda in ws['B']]
    print(columna_b)

# lectura de archivos Excel utilizando pandas
def read_excel_pandas(archivo):
    df = pd.read_excel(archivo, names=columnas)
    
    pd.set_option('display.max_columns', None)  # Mostrar todas las columnas
    pd.set_option('display.width', 1000)       # Aumentar ancho para no saltar líneas

    print("\n--- Vista previa de los datos ---")
    print(df.head()) 
    
    return df

# Extracción de rutas relativas
def extract_relative_paths():
    url_project = os.path.dirname(os.path.abspath(__file__))
    url_excel = os.path.join(url_project, URL_PATH_EXCEL)
    return url_excel

if __name__ == '__main__':
    url_excel = extract_relative_paths()
    #read_xlss_excel(url_excel)
    read_excel_pandas(url_excel)